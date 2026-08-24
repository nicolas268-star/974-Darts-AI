from __future__ import annotations

import csv
import hashlib
import io
import mimetypes
import zipfile
from dataclasses import asdict, dataclass, field
from enum import Enum
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


# Colonnes minimales actuellement attendues par app/parser.py.
NAKKA_REQUIRED_COLUMNS: tuple[str, ...] = (
    "Saison",
    "Jour",
    "Rencontre",
    "Match",
    "S/D",
    "Team",
    "Joueur",
    "Leg",
    "Score",
)

# Feuille utilisée par l'import historique du projet.
PREFERRED_NAKKA_SHEET = "PvP"

# Limites défensives : l'analyse ne doit pas charger un fichier démesuré.
MAX_UPLOAD_BYTES = 50 * 1024 * 1024  # 50 Mo
MAX_HEADER_SCAN_ROWS = 25
MAX_CSV_SAMPLE_BYTES = 256 * 1024


class FileKind(str, Enum):
    XLSX = "xlsx"
    XLSM = "xlsm"
    CSV = "csv"
    UNKNOWN = "unknown"


class DetectionStatus(str, Enum):
    READY = "READY"
    CHECK = "CHECK"
    BLOCKED = "BLOCKED"


@dataclass(frozen=True)
class DetectionIssue:
    code: str
    severity: str
    message: str
    field: str | None = None

    def to_dict(self) -> dict[str, str | None]:
        return asdict(self)


@dataclass
class DetectedImportFile:
    filename: str
    sha256: str
    size_bytes: int
    kind: FileKind
    mime_type: str | None
    status: DetectionStatus

    sheet_name: str | None = None
    available_sheets: list[str] = field(default_factory=list)
    header_row: int | None = None
    columns: list[str] = field(default_factory=list)
    missing_required_columns: list[str] = field(default_factory=list)

    csv_encoding: str | None = None
    csv_delimiter: str | None = None

    is_probable_nakka_export: bool = False
    issues: list[DetectionIssue] = field(default_factory=list)

    def to_dict(self) -> dict[str, object]:
        return {
            "filename": self.filename,
            "sha256": self.sha256,
            "sizeBytes": self.size_bytes,
            "kind": self.kind.value,
            "mimeType": self.mime_type,
            "status": self.status.value,
            "sheetName": self.sheet_name,
            "availableSheets": self.available_sheets,
            "headerRow": self.header_row,
            "columns": self.columns,
            "missingRequiredColumns": self.missing_required_columns,
            "csvEncoding": self.csv_encoding,
            "csvDelimiter": self.csv_delimiter,
            "isProbableNakkaExport": self.is_probable_nakka_export,
            "issues": [issue.to_dict() for issue in self.issues],
        }


class FileDetectionError(ValueError):
    """Erreur contrôlée pendant l'identification d'un fichier d'import."""


def _clean_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalized_headers(values: Iterable[object]) -> list[str]:
    return [_clean_header(value) for value in values]


def _required_column_matches(columns: Iterable[str]) -> tuple[list[str], list[str]]:
    available = {column.casefold(): column for column in columns if column}
    missing = [
        required
        for required in NAKKA_REQUIRED_COLUMNS
        if required.casefold() not in available
    ]
    matched = [
        required
        for required in NAKKA_REQUIRED_COLUMNS
        if required.casefold() in available
    ]
    return matched, missing


def _status_from_issues(issues: Iterable[DetectionIssue]) -> DetectionStatus:
    severities = {issue.severity.upper() for issue in issues}
    if "CRITICAL" in severities:
        return DetectionStatus.BLOCKED
    if "WARNING" in severities:
        return DetectionStatus.CHECK
    return DetectionStatus.READY


def _detect_kind(content: bytes, filename: str) -> FileKind:
    suffix = Path(filename).suffix.casefold()

    # XLSX/XLSM sont des conteneurs ZIP. On vérifie aussi la signature pour ne
    # pas se fier uniquement à l'extension fournie par le navigateur.
    if content.startswith(b"PK\x03\x04"):
        if suffix == ".xlsm":
            return FileKind.XLSM
        return FileKind.XLSX

    if suffix == ".csv":
        return FileKind.CSV

    # Certains exports CSV arrivent avec une extension incorrecte.
    if b"\x00" not in content[:4096]:
        sample = content[:4096]
        if any(separator in sample for separator in (b";", b",", b"\t")):
            return FileKind.CSV

    return FileKind.UNKNOWN


def _find_excel_header(
    worksheet,
    required_columns: tuple[str, ...] = NAKKA_REQUIRED_COLUMNS,
) -> tuple[int | None, list[str]]:
    best_row: int | None = None
    best_headers: list[str] = []
    best_score = -1

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(worksheet.max_row, MAX_HEADER_SCAN_ROWS),
            values_only=True,
        ),
        start=1,
    ):
        headers = _normalized_headers(row)
        matched, _ = _required_column_matches(headers)
        score = len(matched)

        if score > best_score:
            best_row = row_number
            best_headers = headers
            best_score = score

        if score == len(required_columns):
            break

    return best_row, [header for header in best_headers if header]


def _detect_excel(content: bytes, filename: str, kind: FileKind) -> DetectedImportFile:
    issues: list[DetectionIssue] = []

    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_vba=(kind == FileKind.XLSM),
        )
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        issues.append(
            DetectionIssue(
                code="FILE-EXCEL-INVALID",
                severity="CRITICAL",
                message=f"Le classeur Excel est illisible : {exc}",
            )
        )
        return DetectedImportFile(
            filename=filename,
            sha256=hashlib.sha256(content).hexdigest(),
            size_bytes=len(content),
            kind=kind,
            mime_type=mimetypes.guess_type(filename)[0],
            status=DetectionStatus.BLOCKED,
            issues=issues,
        )

    available_sheets = list(workbook.sheetnames)

    if PREFERRED_NAKKA_SHEET in available_sheets:
        selected_sheet = PREFERRED_NAKKA_SHEET
    elif available_sheets:
        selected_sheet = available_sheets[0]
        issues.append(
            DetectionIssue(
                code="FILE-SHEET-FALLBACK",
                severity="WARNING",
                field="sheet",
                message=(
                    f"La feuille '{PREFERRED_NAKKA_SHEET}' est absente. "
                    f"La feuille '{selected_sheet}' a été analysée à la place."
                ),
            )
        )
    else:
        selected_sheet = None
        issues.append(
            DetectionIssue(
                code="FILE-NO-SHEET",
                severity="CRITICAL",
                field="sheet",
                message="Le classeur ne contient aucune feuille.",
            )
        )

    header_row: int | None = None
    columns: list[str] = []
    missing = list(NAKKA_REQUIRED_COLUMNS)

    if selected_sheet is not None:
        worksheet = workbook[selected_sheet]
        header_row, columns = _find_excel_header(worksheet)
        _, missing = _required_column_matches(columns)

        if header_row is None or not columns:
            issues.append(
                DetectionIssue(
                    code="FILE-NO-HEADER",
                    severity="CRITICAL",
                    field="columns",
                    message="Aucune ligne d'en-tête exploitable n'a été détectée.",
                )
            )
        elif missing:
            issues.append(
                DetectionIssue(
                    code="FILE-MISSING-COLUMNS",
                    severity="CRITICAL",
                    field="columns",
                    message="Colonnes Nakka obligatoires absentes : " + ", ".join(missing),
                )
            )

    probable_nakka = not missing and selected_sheet is not None
    status = _status_from_issues(issues)

    workbook.close()

    return DetectedImportFile(
        filename=filename,
        sha256=hashlib.sha256(content).hexdigest(),
        size_bytes=len(content),
        kind=kind,
        mime_type=mimetypes.guess_type(filename)[0],
        status=status,
        sheet_name=selected_sheet,
        available_sheets=available_sheets,
        header_row=header_row,
        columns=columns,
        missing_required_columns=missing,
        is_probable_nakka_export=probable_nakka,
        issues=issues,
    )


def _decode_csv(content: bytes) -> tuple[str, str]:
    encodings = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
    for encoding in encodings:
        try:
            return content.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise FileDetectionError("Impossible de déterminer l'encodage du fichier CSV.")


def _detect_csv(content: bytes, filename: str) -> DetectedImportFile:
    issues: list[DetectionIssue] = []

    try:
        decoded, encoding = _decode_csv(content[:MAX_CSV_SAMPLE_BYTES])
    except FileDetectionError as exc:
        issues.append(
            DetectionIssue(
                code="FILE-CSV-ENCODING",
                severity="CRITICAL",
                message=str(exc),
            )
        )
        return DetectedImportFile(
            filename=filename,
            sha256=hashlib.sha256(content).hexdigest(),
            size_bytes=len(content),
            kind=FileKind.CSV,
            mime_type=mimetypes.guess_type(filename)[0] or "text/csv",
            status=DetectionStatus.BLOCKED,
            issues=issues,
        )

    try:
        dialect = csv.Sniffer().sniff(decoded, delimiters=";,\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";"
        issues.append(
            DetectionIssue(
                code="FILE-CSV-DELIMITER",
                severity="WARNING",
                field="delimiter",
                message="Séparateur CSV non déterminé avec certitude ; ';' est utilisé.",
            )
        )

    reader = csv.reader(io.StringIO(decoded), delimiter=delimiter)
    rows = []
    for _, row in zip(range(MAX_HEADER_SCAN_ROWS), reader):
        rows.append(row)

    best_row: int | None = None
    best_headers: list[str] = []
    best_score = -1

    for row_number, row in enumerate(rows, start=1):
        headers = _normalized_headers(row)
        matched, _ = _required_column_matches(headers)
        if len(matched) > best_score:
            best_row = row_number
            best_headers = headers
            best_score = len(matched)

    columns = [header for header in best_headers if header]
    _, missing = _required_column_matches(columns)

    if not columns:
        issues.append(
            DetectionIssue(
                code="FILE-NO-HEADER",
                severity="CRITICAL",
                field="columns",
                message="Aucune ligne d'en-tête exploitable n'a été détectée.",
            )
        )
    elif missing:
        issues.append(
            DetectionIssue(
                code="FILE-MISSING-COLUMNS",
                severity="CRITICAL",
                field="columns",
                message="Colonnes Nakka obligatoires absentes : " + ", ".join(missing),
            )
        )

    return DetectedImportFile(
        filename=filename,
        sha256=hashlib.sha256(content).hexdigest(),
        size_bytes=len(content),
        kind=FileKind.CSV,
        mime_type=mimetypes.guess_type(filename)[0] or "text/csv",
        status=_status_from_issues(issues),
        header_row=best_row,
        columns=columns,
        missing_required_columns=missing,
        csv_encoding=encoding,
        csv_delimiter=delimiter,
        is_probable_nakka_export=not missing,
        issues=issues,
    )


def detect_import_file(content: bytes, filename: str) -> DetectedImportFile:
    """
    Identifie un fichier avant son passage dans le parser Nakka.

    Cette fonction ne publie aucune donnée et ne contacte pas Supabase.
    Elle vérifie uniquement :
      - le type réel du fichier ;
      - sa taille et son empreinte SHA-256 ;
      - la feuille Excel utilisée ;
      - la ligne d'en-tête ;
      - la présence des colonnes minimales attendues.

    Args:
        content: contenu binaire complet envoyé par FastAPI.
        filename: nom d'origine du fichier.

    Returns:
        DetectedImportFile: résultat structuré et sérialisable avec ``to_dict``.

    Raises:
        FileDetectionError: lorsque l'entrée elle-même est invalide.
    """

    safe_filename = Path(filename or "upload").name

    if not content:
        raise FileDetectionError("Le fichier envoyé est vide.")

    if len(content) > MAX_UPLOAD_BYTES:
        raise FileDetectionError(
            f"Le fichier dépasse la taille maximale autorisée de "
            f"{MAX_UPLOAD_BYTES // (1024 * 1024)} Mo."
        )

    kind = _detect_kind(content, safe_filename)

    if kind in (FileKind.XLSX, FileKind.XLSM):
        return _detect_excel(content, safe_filename, kind)

    if kind == FileKind.CSV:
        return _detect_csv(content, safe_filename)

    issue = DetectionIssue(
        code="FILE-UNSUPPORTED",
        severity="CRITICAL",
        message=(
            "Format non pris en charge. Formats acceptés : "
            ".xlsx, .xlsm et .csv."
        ),
    )
    return DetectedImportFile(
        filename=safe_filename,
        sha256=hashlib.sha256(content).hexdigest(),
        size_bytes=len(content),
        kind=FileKind.UNKNOWN,
        mime_type=mimetypes.guess_type(safe_filename)[0],
        status=DetectionStatus.BLOCKED,
        issues=[issue],
    )
