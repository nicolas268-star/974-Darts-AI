from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
import math
import re
import unicodedata
from typing import Any

from openpyxl import load_workbook
from supabase import Client


FIRST9_SHEET_NAMES = (
    "Nakka_Player_Raw",
    "Nakka Player Raw",
    "Nakka Players",
)


class First9SourceError(ValueError):
    """Raised when the official Nakka First 9 source cannot be read safely."""


def _rows(response: Any) -> list[dict[str, Any]]:
    return list(getattr(response, "data", None) or [])


def _normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", text)


def _exact_key(value: Any) -> str:
    text = unicodedata.normalize("NFC", str(value or "").strip())
    return re.sub(r"\s+", " ", text).casefold()


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def _integer(value: Any) -> int:
    parsed = _number(value)
    return max(0, int(parsed)) if parsed is not None else 0


def _find_header(headers: list[Any], candidates: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if _normalize(header) in candidates:
            return index
    return None


@dataclass(frozen=True)
class First9SourceRow:
    player: str
    team: str
    first_9: float
    total_darts: int
    source_row: int


@dataclass(frozen=True)
class First9Workbook:
    sheet_name: str
    rows_seen: int
    rows: list[First9SourceRow]
    rejected_rows: list[dict[str, Any]]


def parse_first9_workbook(content: bytes) -> First9Workbook:
    """Read official season First 9 values from the Nakka player summary."""
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise First9SourceError(
            f"Classeur First 9 illisible : {type(exc).__name__}"
        ) from exc

    normalized_sheet_names = {
        _normalize(name): name
        for name in workbook.sheetnames
    }
    sheet_name = next(
        (
            normalized_sheet_names[_normalize(candidate)]
            for candidate in FIRST9_SHEET_NAMES
            if _normalize(candidate) in normalized_sheet_names
        ),
        None,
    )
    if sheet_name is None:
        raise First9SourceError(
            "La feuille Nakka_Player_Raw est absente du classeur."
        )

    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    headers = list(next(iterator, ()))
    if not headers:
        raise First9SourceError("La feuille Nakka_Player_Raw est vide.")

    player_index = _find_header(headers, {"player", "joueur"})
    team_index = _find_header(headers, {"team", "equipe"})
    first9_index = _find_header(
        headers,
        {
            "first9",
            "first9average",
            "first9avaerage",
            "avgfirst9darts",
            "averagefirst9darts",
        },
    )
    first9_score_index = _find_header(
        headers,
        {"first9score", "scorefirst9"},
    )
    first9_darts_index = _find_header(
        headers,
        {"first9darts", "dartsfirst9"},
    )
    total_darts_index = _find_header(
        headers,
        {"totaldarts", "dartstotal", "flecheslancees"},
    )

    if player_index is None or team_index is None:
        raise First9SourceError(
            "Colonnes Player/Team absentes de Nakka_Player_Raw."
        )
    if first9_index is None and (
        first9_score_index is None or first9_darts_index is None
    ):
        raise First9SourceError(
            "Colonne First 9 absente de Nakka_Player_Raw."
        )

    selected: dict[tuple[str, str], First9SourceRow] = {}
    rejected_rows: list[dict[str, Any]] = []
    rows_seen = 0

    for source_row, values in enumerate(iterator, start=2):
        rows_seen += 1
        row = list(values)
        player = str(row[player_index] or "").strip()
        team = str(row[team_index] or "").strip()
        if not player or not team:
            continue

        first_9 = (
            _number(row[first9_index])
            if first9_index is not None
            else None
        )
        if first_9 is None:
            first9_score = (
                _number(row[first9_score_index])
                if first9_score_index is not None
                else None
            )
            first9_darts = (
                _number(row[first9_darts_index])
                if first9_darts_index is not None
                else None
            )
            if (
                first9_score is not None
                and first9_darts is not None
                and first9_darts > 0
            ):
                first_9 = first9_score / first9_darts * 3

        if first_9 is None or not 0 <= first_9 <= 180:
            rejected_rows.append({
                "row": source_row,
                "player": player,
                "team": team,
                "reason": "FIRST9_INVALID",
            })
            continue

        total_darts = (
            _integer(row[total_darts_index])
            if total_darts_index is not None
            else 0
        )
        candidate = First9SourceRow(
            player=player,
            team=team,
            first_9=round(first_9, 2),
            total_darts=total_darts,
            source_row=source_row,
        )
        key = (_normalize(player), _normalize(team))
        previous = selected.get(key)
        if previous is None or (
            candidate.total_darts,
            candidate.source_row,
        ) > (
            previous.total_darts,
            previous.source_row,
        ):
            selected[key] = candidate

    return First9Workbook(
        sheet_name=sheet_name,
        rows_seen=rows_seen,
        rows=list(selected.values()),
        rejected_rows=rejected_rows,
    )


class First9ProfileSyncService:
    """Synchronize official season First 9 averages into player profiles."""

    def __init__(self, db: Client):
        self.db = db

    def _active_season(self) -> dict[str, Any]:
        active = _rows(
            self.db.table("seasons")
            .select("id,name,is_active")
            .eq("is_active", True)
            .limit(1)
            .execute()
        )
        if active:
            return active[0]

        seasons = _rows(
            self.db.table("seasons")
            .select("id,name,is_active")
            .order("name", desc=True)
            .limit(1)
            .execute()
        )
        if not seasons:
            raise First9SourceError(
                "Aucune saison n'est disponible pour le First 9."
            )
        return seasons[0]

    def _canonical_player_map(
        self,
        players: list[dict[str, Any]],
    ) -> dict[str, str]:
        direct = {
            str(player["id"]): str(player["id"])
            for player in players
            if player.get("id")
        }
        try:
            identities = _rows(
                self.db.table("player_identities")
                .select(
                    "id,canonical_player_id,status,"
                    "merged_into_identity_id"
                )
                .execute()
            )
            aliases = _rows(
                self.db.table("player_aliases")
                .select("identity_id,source_player_id")
                .execute()
            )
        except Exception:
            return direct

        identity_by_id = {
            str(row["id"]): row
            for row in identities
            if row.get("id")
        }
        identity_for_player: dict[str, str] = {}
        for row in identities:
            if row.get("canonical_player_id") and row.get("id"):
                identity_for_player[str(row["canonical_player_id"])] = str(
                    row["id"]
                )
        for row in aliases:
            if row.get("source_player_id") and row.get("identity_id"):
                identity_for_player[str(row["source_player_id"])] = str(
                    row["identity_id"]
                )

        def resolve(player_id: str) -> str:
            identity_id = identity_for_player.get(player_id)
            visited: set[str] = set()
            while identity_id and identity_id not in visited:
                visited.add(identity_id)
                identity = identity_by_id.get(identity_id)
                if not identity:
                    break
                if (
                    identity.get("status") == "MERGED"
                    and identity.get("merged_into_identity_id")
                ):
                    identity_id = str(
                        identity["merged_into_identity_id"]
                    )
                    continue
                return str(
                    identity.get("canonical_player_id") or player_id
                )
            return player_id

        return {
            player_id: resolve(player_id)
            for player_id in direct
        }

    def sync_workbook(
        self,
        content: bytes,
        filename: str,
    ) -> dict[str, Any]:
        parsed = parse_first9_workbook(content)
        season = self._active_season()

        teams = _rows(
            self.db.table("teams")
            .select("id,name")
            .execute()
        )
        players = _rows(
            self.db.table("players")
            .select("id,display_name,team_id")
            .execute()
        )
        team_name_by_id = {
            str(team["id"]): str(team.get("name") or "")
            for team in teams
            if team.get("id")
        }
        player_by_name_team: dict[
            tuple[str, str],
            list[dict[str, Any]],
        ] = {}
        player_by_exact_name_team: dict[
            tuple[str, str],
            list[dict[str, Any]],
        ] = {}
        players_by_name: dict[str, list[dict[str, Any]]] = {}
        for player in players:
            player_id = player.get("id")
            if not player_id:
                continue
            name_key = _normalize(player.get("display_name"))
            team_key = _normalize(
                team_name_by_id.get(str(player.get("team_id")))
            )
            player_by_name_team.setdefault(
                (name_key, team_key),
                [],
            ).append(player)
            player_by_exact_name_team.setdefault(
                (
                    _exact_key(player.get("display_name")),
                    _exact_key(
                        team_name_by_id.get(str(player.get("team_id")))
                    ),
                ),
                [],
            ).append(player)
            players_by_name.setdefault(name_key, []).append(player)

        canonical_by_player = self._canonical_player_map(players)
        members_by_canonical: dict[str, set[str]] = defaultdict(set)
        for source_player_id, canonical_player_id in (
            canonical_by_player.items()
        ):
            members_by_canonical[canonical_player_id].add(
                source_player_id
            )
            members_by_canonical[canonical_player_id].add(
                canonical_player_id
            )

        matched: dict[str, tuple[First9SourceRow, str]] = {}
        unmatched: list[dict[str, Any]] = []
        ambiguous: list[dict[str, Any]] = []

        for source in parsed.rows:
            name_key = _normalize(source.player)
            team_key = _normalize(source.team)
            candidates = player_by_exact_name_team.get(
                (
                    _exact_key(source.player),
                    _exact_key(source.team),
                ),
                [],
            )
            match_type = "EXACT_NAME_AND_TEAM"
            if not candidates:
                candidates = player_by_name_team.get(
                    (name_key, team_key),
                    [],
                )
                match_type = "NORMALIZED_NAME_AND_TEAM"
            if not candidates:
                unique_name_candidates = players_by_name.get(name_key, [])
                if len(unique_name_candidates) == 1:
                    candidates = unique_name_candidates
                    match_type = "UNIQUE_NAME"

            if not candidates:
                unmatched.append({
                    "row": source.source_row,
                    "player": source.player,
                    "team": source.team,
                })
                continue
            if len(candidates) > 1:
                ambiguous.append({
                    "row": source.source_row,
                    "player": source.player,
                    "team": source.team,
                    "candidate_count": len(candidates),
                })
                continue

            source_player_id = str(candidates[0]["id"])
            canonical_player_id = canonical_by_player.get(
                source_player_id,
                source_player_id,
            )
            previous = matched.get(canonical_player_id)
            if previous is None or (
                source.total_darts,
                source.source_row,
            ) > (
                previous[0].total_darts,
                previous[0].source_row,
            ):
                matched[canonical_player_id] = (source, match_type)

        updated_at = datetime.now(timezone.utc).isoformat()
        payload: list[dict[str, Any]] = []
        for canonical_player_id, (source, _match_type) in matched.items():
            target_player_ids = members_by_canonical.get(
                canonical_player_id,
                {canonical_player_id},
            )
            for player_id in sorted(target_player_ids):
                payload.append({
                    "player_id": player_id,
                    "season_id": season["id"],
                    "first_9": source.first_9,
                    "updated_at": updated_at,
                })
        if payload:
            self.db.table("player_profiles").upsert(
                payload,
                on_conflict="player_id,season_id",
            ).execute()

        return {
            "status": "SYNCED",
            "filename": filename,
            "sheet": parsed.sheet_name,
            "season": {
                "id": season.get("id"),
                "name": season.get("name"),
            },
            "rows_seen": parsed.rows_seen,
            "valid_source_rows": len(parsed.rows),
            "identities_matched": len(matched),
            "profiles_updated": len(payload),
            "rejected_rows": parsed.rejected_rows,
            "unmatched_rows": unmatched,
            "ambiguous_rows": ambiguous,
            "source": "NAKKA_PLAYER_RAW",
            "calculation": (
                "Valeur officielle de la feuille Nakka. "
                "Si first9Score/first9Darts sont présents : "
                "first9Score / first9Darts × 3."
            ),
        }
