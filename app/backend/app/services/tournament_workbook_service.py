from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timezone
from io import BytesIO
from html import unescape
import hashlib
import json
import math
from pathlib import Path
import re
import unicodedata
from typing import Any
from urllib.request import Request, urlopen

from openpyxl import load_workbook


TOURNAMENT_CODES = {"T1", "T2"}
KNOWN_TOURNAMENT_METADATA = {
    "t_lJst_9313": {
        "date": "2026-05-14",
        "event_name": "Tournoi Papangue Darts Club 974",
    },
}
TOURNAMENT_CACHE = (
    Path(__file__).resolve().parents[2]
    / "data"
    / "sprint14_tournaments.json"
)
FRENCH_MONTHS = {
    "janvier": 1,
    "fevrier": 2,
    "mars": 3,
    "avril": 4,
    "mai": 5,
    "juin": 6,
    "juillet": 7,
    "aout": 8,
    "septembre": 9,
    "octobre": 10,
    "novembre": 11,
    "decembre": 12,
}
FRENCH_MONTH_LABELS = (
    "",
    "janvier",
    "février",
    "mars",
    "avril",
    "mai",
    "juin",
    "juillet",
    "août",
    "septembre",
    "octobre",
    "novembre",
    "décembre",
)


class TournamentWorkbookError(ValueError):
    """Raised when tournament rows cannot be read safely."""


def _normalize(value: Any) -> str:
    text = unicodedata.normalize(
        "NFKD",
        str(value or "").strip().lower(),
    )
    text = "".join(
        char
        for char in text
        if not unicodedata.combining(char)
    )
    return re.sub(r"[^a-z0-9]+", "", text)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def _integer(value: Any) -> int:
    parsed = _number(value)
    return int(parsed) if parsed is not None else 0


def _header_index(headers: list[Any], names: set[str]) -> int | None:
    expected = {_normalize(name) for name in names}
    for index, header in enumerate(headers):
        if _normalize(header) in expected:
            return index
    return None


def _row_value(
    row: list[Any],
    indexes: dict[str, int | None],
    key: str,
) -> Any:
    index = indexes.get(key)
    return row[index] if index is not None and index < len(row) else None


def _side_key(value: str) -> str:
    words = re.findall(
        r"[a-z0-9]+",
        unicodedata.normalize("NFKD", value.lower()),
    )
    words = [
        word
        for word in words
        if word not in {"le", "la", "les", "l"}
    ]
    compact = "".join(words)
    if compact.endswith("s") and len(compact) > 4:
        compact = compact[:-1]
    return compact


def _split_encounter(value: str) -> tuple[str, str]:
    parts = re.split(r"\s+vs\s+", value, maxsplit=1, flags=re.I)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return value.strip(), "Adversaire"


def _team_side(team: str, home: str, away: str) -> int | None:
    team_key = _side_key(team)
    home_key = _side_key(home)
    away_key = _side_key(away)
    if not team_key:
        return None
    if (
        team_key == home_key
        or team_key in home_key
        or home_key in team_key
    ):
        return 0
    if (
        team_key == away_key
        or team_key in away_key
        or away_key in team_key
    ):
        return 1
    return None


def _stable_id(*parts: Any) -> str:
    raw = "|".join(_text(part) for part in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def _source_phase(
    source_url: str,
) -> tuple[str | None, str | None, int | None]:
    match = re.search(
        r"[?&]tmid=(.+?)_(rr|t)_(\d+)_",
        source_url,
        flags=re.I,
    )
    if not match:
        return None, None, None
    return (
        match.group(1),
        match.group(2).lower(),
        int(match.group(3)),
    )


def _date_iso(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    if not text:
        return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    normalized = unicodedata.normalize("NFKD", text.lower())
    normalized = "".join(
        char
        for char in normalized
        if not unicodedata.combining(char)
    )
    match = re.search(
        r"\b(\d{1,2})\s+([a-z]+)\s+(20\d{2})\b",
        normalized,
    )
    if not match:
        return None
    month = FRENCH_MONTHS.get(match.group(2))
    if month is None:
        return None
    try:
        return date(
            int(match.group(3)),
            month,
            int(match.group(1)),
        ).isoformat()
    except ValueError:
        return None


def _date_label(value: str | None) -> str | None:
    if not value:
        return None
    try:
        parsed = date.fromisoformat(value)
    except ValueError:
        return None
    return (
        f"{parsed.day} "
        f"{FRENCH_MONTH_LABELS[parsed.month]} "
        f"{parsed.year}"
    )


def _knockout_label(index: int, maximum: int) -> str:
    distance = maximum - index
    labels = {
        0: "Finale",
        1: "Demi-finales",
        2: "Quarts de finale",
        3: "1/8 de finale",
        4: "1/16 de finale",
    }
    return labels.get(distance, f"Tour éliminatoire {index + 1}")


def _apply_stage_labels(matches: list[dict[str, Any]]) -> None:
    knockout_indexes = [
        int(match["stage_index"])
        for match in matches
        if match.get("phase") == "KNOCKOUT"
        and match.get("stage_index") is not None
    ]
    maximum = max(knockout_indexes, default=0)
    for match in matches:
        index = match.get("stage_index")
        if match.get("phase") == "POOL" and index is not None:
            match["stage_label"] = f"Poule {int(index)}"
        elif match.get("phase") == "KNOCKOUT" and index is not None:
            match["stage_label"] = _knockout_label(
                int(index),
                maximum,
            )
        else:
            match["stage_label"] = "Phase non identifiée"


def _organize_stages(
    matches: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    pool_groups: dict[int, list[dict[str, Any]]] = defaultdict(list)
    knockout_groups: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for match in matches:
        index = match.get("stage_index")
        if index is None:
            continue
        if match.get("phase") == "POOL":
            pool_groups[int(index)].append(match)
        elif match.get("phase") == "KNOCKOUT":
            knockout_groups[int(index)].append(match)

    pools = [
        {
            "code": f"rr_{index}",
            "name": f"Poule {index}",
            "order": index,
            "matches": group,
        }
        for index, group in sorted(pool_groups.items())
    ]
    bracket = [
        {
            "code": f"t_{index}",
            "name": group[0].get("stage_label"),
            "order": index,
            "matches": group,
        }
        for index, group in sorted(knockout_groups.items())
    ]
    return pools, bracket


def _tournament_metadata(
    rows: list[dict[str, Any]],
    code: str,
) -> dict[str, Any]:
    tournament_ids = sorted({
        str(row.get("source_tournament_id"))
        for row in rows
        if row.get("source_tournament_id")
    })
    source_id = tournament_ids[0] if tournament_ids else None
    raw_dates = [
        _date_iso(row.get("tournament_date"))
        for row in rows
    ]
    parsed_date = next(
        (value for value in raw_dates if value),
        None,
    )
    known = KNOWN_TOURNAMENT_METADATA.get(
        source_id or "",
        {},
    )
    parsed_date = parsed_date or known.get("date")
    event_name = known.get("event_name") or f"Tournoi amical {code}"
    label = _date_label(parsed_date)
    return {
        "source_tournament_id": source_id,
        "date": parsed_date,
        "date_label": label,
        "event_name": event_name,
        "name": label or f"Tournoi amical {code}",
    }


def enrich_tournament_metadata(
    payload: dict[str, Any],
    timeout_seconds: float = 4.0,
) -> dict[str, Any]:
    """Best-effort official N01 title/date lookup; never blocks the import."""

    month_pattern = "|".join(FRENCH_MONTHS)
    for tournament in payload.get("tournaments") or []:
        if tournament.get("date"):
            continue
        source_id = _text(tournament.get("source_tournament_id"))
        if not source_id:
            continue
        url = (
            "https://n01darts.com/n01/tournament/comp.php?id="
            + source_id
        )
        try:
            request = Request(
                url,
                headers={"User-Agent": "974-Darts-AI/14.1"},
            )
            with urlopen(
                request,
                timeout=timeout_seconds,
            ) as response:
                html = response.read(1_000_000).decode(
                    "utf-8",
                    errors="replace",
                )
        except Exception:
            continue

        plain = unescape(re.sub(r"<[^>]+>", " ", html))
        plain = re.sub(r"\s+", " ", plain)
        normalized_plain = unicodedata.normalize(
            "NFKD",
            plain.lower(),
        )
        normalized_plain = "".join(
            char
            for char in normalized_plain
            if not unicodedata.combining(char)
        )
        date_match = re.search(
            rf"\b(\d{{1,2}})\s+({month_pattern})\s+(20\d{{2}})\b",
            normalized_plain,
            flags=re.I,
        )
        if not date_match:
            continue
        parsed_date = _date_iso(date_match.group(0))
        if not parsed_date:
            continue
        event_match = re.search(
            r"(Tournoi[^.]{3,120}?)\s+du\s+\d{1,2}\s+",
            plain,
            flags=re.I,
        )
        tournament["date"] = parsed_date
        tournament["date_label"] = _date_label(parsed_date)
        tournament["name"] = tournament["date_label"]
        if event_match:
            tournament["event_name"] = event_match.group(1).strip()
    return payload


def _weighted_average(rows: list[dict[str, Any]]) -> float | None:
    weighted = [
        (
            float(row["average_3_darts"]),
            int(row.get("darts_thrown") or 0),
        )
        for row in rows
        if row.get("average_3_darts") is not None
        and int(row.get("darts_thrown") or 0) > 0
    ]
    if weighted:
        total_darts = sum(darts for _, darts in weighted)
        return round(
            sum(value * darts for value, darts in weighted)
            / total_darts,
            2,
        )
    values = [
        float(row["average_3_darts"])
        for row in rows
        if row.get("average_3_darts") is not None
    ]
    return round(sum(values) / len(values), 2) if values else None


def _aggregate_participants(
    rows: list[dict[str, Any]],
    group_key: str,
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        label = _text(row.get(group_key))
        if label:
            grouped[label].append(row)

    result: list[dict[str, Any]] = []
    for label, group in grouped.items():
        leg_keys = {
            (
                row["match_id"],
                _text(row.get("leg")),
            )
            for row in group
        }
        won_leg_keys = {
            (
                row["match_id"],
                _text(row.get("leg")),
            )
            for row in group
            if row.get("result") == "V"
        }
        finishes = [
            int(row["finish"])
            for row in group
            if row.get("finish") not in (None, 0)
        ]
        item = {
            "name": label,
            "legs_played": len(leg_keys),
            "legs_won": len(won_leg_keys),
            "average_3_darts": _weighted_average(group),
            "best_finish": max(finishes) if finishes else None,
            "scores_180": sum(
                int(row.get("scores_180") or 0)
                for row in group
            ),
            "scores_140": sum(
                int(row.get("scores_140") or 0)
                for row in group
            ),
            "scores_100": sum(
                int(row.get("scores_100") or 0)
                for row in group
            ),
            "no_score": sum(
                int(row.get("no_score") or 0)
                for row in group
            ),
        }
        if group_key == "player":
            teams = sorted({
                _text(row.get("team"))
                for row in group
                if _text(row.get("team"))
            })
            item["teams"] = teams
            item["team"] = " / ".join(teams)
        else:
            item["players"] = sorted({
                _text(row.get("player"))
                for row in group
                if _text(row.get("player"))
            })
        result.append(item)

    result.sort(
        key=lambda item: (
            -int(item.get("legs_won") or 0),
            -float(item.get("average_3_darts") or 0),
            str(item.get("name") or "").lower(),
        )
    )
    return result


def _build_matches(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row["match_id"])].append(row)

    matches: list[dict[str, Any]] = []
    for match_id, match_rows in grouped.items():
        first = match_rows[0]
        home, away = _split_encounter(first["encounter"])
        source_url = _text(first.get("source_url"))
        source_tournament_id, source_phase, stage_index = (
            _source_phase(source_url)
        )
        by_leg: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in match_rows:
            by_leg[_text(row.get("leg"))].append(row)

        score = [0, 0]
        unresolved = 0
        for leg_rows in by_leg.values():
            candidates: set[int] = set()
            for row in leg_rows:
                side = _team_side(
                    _text(row.get("team")),
                    home,
                    away,
                )
                result = _text(row.get("result")).upper()
                if side is None or result not in {"V", "D"}:
                    continue
                candidates.add(side if result == "V" else 1 - side)
            if len(candidates) == 1:
                score[candidates.pop()] += 1
            else:
                unresolved += 1

        if score[0] > score[1]:
            winner = home
        elif score[1] > score[0]:
            winner = away
        else:
            winner = None

        tracked_teams = sorted({
            _text(row.get("team"))
            for row in match_rows
            if _text(row.get("team"))
        })
        tracked_players = sorted({
            _text(row.get("player"))
            for row in match_rows
            if _text(row.get("player"))
        })
        matches.append({
            "id": match_id,
            "match_number": first.get("match_number"),
            "encounter": first.get("encounter"),
            "mode": first.get("mode"),
            "home": home,
            "away": away,
            "home_score": score[0],
            "away_score": score[1],
            "winner": winner,
            "legs": len(by_leg),
            "unresolved_legs": unresolved,
            "result_complete": unresolved == 0,
            "tracked_teams": tracked_teams,
            "tracked_players": tracked_players,
            "source_url": source_url or None,
            "source_tournament_id": source_tournament_id,
            "phase": (
                "POOL"
                if source_phase == "rr"
                else "KNOCKOUT"
                if source_phase == "t"
                else "UNKNOWN"
            ),
            "stage_code": (
                f"{source_phase}_{stage_index}"
                if source_phase is not None
                and stage_index is not None
                else None
            ),
            "stage_index": stage_index,
            "stage_label": None,
        })

    def match_sort_key(item: dict[str, Any]) -> tuple[int, str]:
        number = _number(item.get("match_number"))
        return (
            int(number) if number is not None else 10**9,
            str(item.get("encounter") or "").lower(),
        )

    matches.sort(key=match_sort_key)
    _apply_stage_labels(matches)
    return matches


def parse_tournament_workbook(
    content: bytes,
    filename: str,
) -> dict[str, Any]:
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise TournamentWorkbookError(
            f"Classeur tournoi illisible : {type(exc).__name__}"
        ) from exc

    if "PvP" not in workbook.sheetnames:
        raise TournamentWorkbookError("La feuille PvP est absente.")
    sheet = workbook["PvP"]
    iterator = sheet.iter_rows(values_only=True)
    headers = list(next(iterator, ()))
    if not headers:
        raise TournamentWorkbookError("La feuille PvP est vide.")

    indexes = {
        "season": _header_index(headers, {"Saison"}),
        "round": _header_index(headers, {"Jour"}),
        "match": _header_index(headers, {"Match"}),
        "nakka_match": _header_index(headers, {"Match Nakka"}),
        "mode": _header_index(headers, {"S/D", "Mode"}),
        "team": _header_index(headers, {"Team", "Équipe"}),
        "player": _header_index(headers, {"Joueur", "Player"}),
        "leg": _header_index(headers, {"Leg"}),
        "score": _header_index(headers, {"Score"}),
        "result": _header_index(headers, {"V/D", "Résultat"}),
        "darts": _header_index(
            headers,
            {"fleches lancees", "flèches lancées"},
        ),
        "average": _header_index(headers, {"Average 3 Darts"}),
        "finish": _header_index(headers, {"Finish"}),
        "scores_180": _header_index(headers, {"180+"}),
        "scores_140": _header_index(headers, {"140+"}),
        "scores_100": _header_index(headers, {"100+"}),
        "no_score": _header_index(headers, {"No Score"}),
        "encounter": _header_index(headers, {"Rencontre"}),
        "source_url": _header_index(
            headers,
            {"Source URL", "URL", "Lien Nakka"},
        ),
        "tournament_date": _header_index(
            headers,
            {"Date", "Date tournoi", "Tournament Date"},
        ),
    }
    required = (
        "season",
        "round",
        "match",
        "mode",
        "team",
        "player",
        "leg",
        "result",
        "encounter",
    )
    missing = [key for key in required if indexes[key] is None]
    if missing:
        raise TournamentWorkbookError(
            "Colonnes tournoi absentes : " + ", ".join(missing)
        )

    rows_by_code: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for source_row, values in enumerate(iterator, start=2):
        row = list(values)
        code = _text(_row_value(row, indexes, "round")).upper()
        if code not in TOURNAMENT_CODES:
            continue
        encounter = _text(_row_value(row, indexes, "encounter"))
        raw_match = _row_value(row, indexes, "nakka_match")
        if raw_match in (None, ""):
            raw_match = _row_value(row, indexes, "match")
        mode = _text(_row_value(row, indexes, "mode"))
        source_url = _text(
            _row_value(row, indexes, "source_url")
        )
        source_tournament_id, _, _ = _source_phase(source_url)
        match_id = _stable_id(
            _row_value(row, indexes, "season"),
            code,
            encounter,
            raw_match,
            mode,
        )
        rows_by_code[code].append({
            "source_row": source_row,
            "season": _text(_row_value(row, indexes, "season")),
            "code": code,
            "match_id": match_id,
            "match_number": raw_match,
            "encounter": encounter,
            "mode": mode,
            "team": _text(_row_value(row, indexes, "team")),
            "player": _text(_row_value(row, indexes, "player")),
            "leg": _row_value(row, indexes, "leg"),
            "score": _number(_row_value(row, indexes, "score")),
            "result": _text(
                _row_value(row, indexes, "result")
            ).upper(),
            "darts_thrown": _integer(
                _row_value(row, indexes, "darts")
            ),
            "average_3_darts": _number(
                _row_value(row, indexes, "average")
            ),
            "finish": _number(
                _row_value(row, indexes, "finish")
            ),
            "scores_180": _integer(
                _row_value(row, indexes, "scores_180")
            ),
            "scores_140": _integer(
                _row_value(row, indexes, "scores_140")
            ),
            "scores_100": _integer(
                _row_value(row, indexes, "scores_100")
            ),
            "no_score": _integer(
                _row_value(row, indexes, "no_score")
            ),
            "source_url": source_url,
            "source_tournament_id": source_tournament_id,
            "tournament_date": _row_value(
                row,
                indexes,
                "tournament_date",
            ),
        })

    tournaments: list[dict[str, Any]] = []
    for code in sorted(rows_by_code):
        rows = rows_by_code[code]
        matches = _build_matches(rows)
        pools, bracket = _organize_stages(matches)
        players = _aggregate_participants(rows, "player")
        duos = _aggregate_participants(rows, "team")
        leg_keys = {
            (row["match_id"], _text(row.get("leg")))
            for row in rows
        }
        seasons = sorted({
            row["season"]
            for row in rows
            if row.get("season")
        })
        complete_results = sum(
            1
            for match in matches
            if match.get("result_complete")
        )
        metadata = _tournament_metadata(rows, code)
        tournaments.append({
            "code": code,
            **metadata,
            "season": seasons[-1] if seasons else None,
            "status": "AVAILABLE",
            "summary": {
                "source_rows": len(rows),
                "matches": len(matches),
                "legs": len(leg_keys),
                "pool_matches": sum(
                    len(group["matches"])
                    for group in pools
                ),
                "knockout_matches": sum(
                    len(round_item["matches"])
                    for round_item in bracket
                ),
                "tracked_players": len(players),
                "tracked_duos": len(duos),
                "complete_results": complete_results,
            },
            "matches": matches,
            "pools": pools,
            "bracket": bracket,
            "players": players,
            "duos": duos,
            "data_quality_notes": [
                (
                    "Les statistiques individuelles couvrent uniquement "
                    "les joueurs présents dans la feuille PvP."
                ),
                (
                    "Les poules et l'arbre affichent uniquement les matchs "
                    "présents dans la feuille PvP ; les branches concernant "
                    "exclusivement des équipes non suivies restent absentes."
                ),
                (
                    "Le First 9 tournoi n'est pas inventé : la feuille PvP "
                    "ne fournit pas cette statistique."
                ),
                (
                    "Les tournois restent exclus du classement et de l'ELO "
                    "officiels du championnat."
                ),
            ],
        })

    return {
        "contract_version": "14.1",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": {
            "filename": filename,
            "sheet": "PvP",
        },
        "tournaments": tournaments,
    }


def save_tournament_cache(
    payload: dict[str, Any],
    path: Path = TOURNAMENT_CACHE,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path


def load_tournament_cache(
    path: Path = TOURNAMENT_CACHE,
) -> dict[str, Any]:
    if not path.is_file():
        return {
            "contract_version": "14.1",
            "source": None,
            "tournaments": [],
        }
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {
            "contract_version": "14.1",
            "source": None,
            "tournaments": [],
        }


def available_tournament_codes(
    payload: dict[str, Any] | None = None,
) -> set[str]:
    """Return built-in and safely imported friendly tournament codes."""
    cache = payload if payload is not None else load_tournament_cache()
    codes = set(TOURNAMENT_CODES)
    for tournament in cache.get("tournaments") or []:
        code = str(tournament.get("code") or "").upper()
        if re.fullmatch(r"T\d+", code):
            codes.add(code)
    return codes
